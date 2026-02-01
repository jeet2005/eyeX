"""
Excel Export Service for Attendance Reports
Generates daily and monthly attendance reports with color-coded status using openpyxl.
"""
import io
from datetime import date, datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Colors
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light Red
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light Yellow
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
YELLOW_FONT = Font(color="9C6500")

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

def generate_daily_excel(data: Dict) -> io.BytesIO:
    """Generate daily attendance Excel with color formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    
    # 1. Header Info
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Daily Attendance Report - {data['date']}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Stats
    ws['A3'] = f"Enrolled: {data['enrolled']}"
    ws['B3'] = f"Present: {data['present']}"
    ws['C3'] = f"Late: {data['late']}"
    ws['D3'] = f"Absent: {data['absent']}"
    
    # 2. Table Headers
    headers = ["Name", "Roll Number", "Arrival Time", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # 3. Data Rows
    row_num = 6
    for student in data.get("students", []):
        name = student.get("name", "")
        roll = student.get("roll_number", "")
        time = student.get("arrival_time", "N/A")
        status = student.get("status", "absent")
        
        ws.cell(row=row_num, column=1, value=name).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=roll).border = THIN_BORDER
        ws.cell(row=row_num, column=3, value=time).border = THIN_BORDER
        
        status_cell = ws.cell(row=row_num, column=4, value=status.title())
        status_cell.border = THIN_BORDER
        
        # Apply Colors
        if status == "present" or status == "on_time":
            status_cell.fill = GREEN_FILL
            status_cell.font = GREEN_FONT
        elif status == "late":
            status_cell.fill = YELLOW_FILL
            status_cell.font = YELLOW_FONT
        elif status == "absent":
            status_cell.fill = RED_FILL
            status_cell.font = RED_FONT
            
        row_num += 1
        
    # Auto-width
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_monthly_excel(data: Dict) -> io.BytesIO:
    """Generate monthly attendance Excel grid with colors."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"
    
    year = data.get("year", datetime.now().year)
    month = data.get("month", datetime.now().month)
    num_days = data.get("num_days", 31)
    
    # 1. Header
    ws['A1'] = f"Monthly Attendance Report - {month}/{year}"
    ws['A1'].font = Font(size=14, bold=True)
    
    ws['A2'] = "Legend: P = Present (Green), L = Late (Yellow), A = Absent (Red)"
    
    # 2. Table Headers
    # Static cols
    ws.cell(row=4, column=1, value="Name").font = Font(bold=True)
    ws.cell(row=4, column=2, value="Roll Number").font = Font(bold=True)
    
    # Days cols
    for day in range(1, num_days + 1):
        cell = ws.cell(row=4, column=2 + day, value=day)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 5 # Narrow columns for days
        
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    
    # 3. Data Rows
    row_num = 5
    for student in data.get("students", []):
        ws.cell(row=row_num, column=1, value=student.get("name", "")).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=student.get("roll_number", "")).border = THIN_BORDER
        
        days = student.get("days", {})
        for day in range(1, num_days + 1):
            day_data = days.get(day, {"status": "absent", "time": "N/A"})
            status = day_data.get("status", "absent")
            time_str = day_data.get("time", "")
            
            cell_val = "A"
            cell_fill = RED_FILL
            cell_font = RED_FONT
            
            if status == "present":
                cell_val = "P"
                cell_fill = GREEN_FILL
                cell_font = GREEN_FONT
            elif status == "late":
                cell_val = "L"
                cell_fill = YELLOW_FILL
                cell_font = YELLOW_FONT
            elif status == "upcoming":  # Future dates (if logic supported it)
                 cell_val = ""
                 cell_fill = WHITE_FILL
                 
            # Add time tooltip/comment if needed, effectively we just show code
            # Or if user wants detailed view, we could put time in cell, but grid is small.
            # Keeping it simple P/L/A.
            
            cell = ws.cell(row=row_num, column=2 + day, value=cell_val)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = cell_fill
            cell.font = cell_font
            cell.border = THIN_BORDER
            
        row_num += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
